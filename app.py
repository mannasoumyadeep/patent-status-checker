import os
import time
import streamlit as st
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoAlertPresentException, TimeoutException
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook, Workbook
from datetime import datetime
import json

# Constants
APPLICATION_STATUS_URL = "https://iprsearch.ipindia.gov.in/PublicSearch/PublicationSearch/ApplicationStatus"
CAPTCHA_URL = "https://iprsearch.ipindia.gov.in/PublicSearch/Captcha/CaptchaAudio"
MAX_RETRIES = 3
RETRY_DELAY = 5  # seconds

# Configure Streamlit page
st.set_page_config(
    page_title="Patent Application Status Checker",
    page_icon="📋",
    layout="wide"
)

def parse_date(date_string):
    """Parse date string to consistent format."""
    if not date_string:
        return None
    try:
        return datetime.strptime(date_string, "%d/%m/%Y").date().strftime("%d/%m/%Y")
    except ValueError:
        return date_string

class ApplicationService:
    def __init__(self):
        self.processed_data = {}
        self.error_applications = []

    def setup_driver(self):
        """Set up Chrome driver with appropriate options."""
        options = Options()
        options.add_argument("--headless=new")
        options.add_argument("--disable-extensions")
        options.add_argument("--disable-gpu")
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        options.page_load_strategy = 'eager'
        
        try:
            # Use ChromeDriverManager to handle driver installation
            service = ChromeService(ChromeDriverManager().install())
            driver = webdriver.Chrome(service=service, options=options)
            wait = WebDriverWait(driver, 15)
            return driver, wait
        except Exception as e:
            st.error(f"Error setting up Chrome driver: {str(e)}")
            raise

    def process_application_number(self, application_number, retry_count=0):
        """Process a single application number and retrieve its status."""
        driver = None
        try:
            driver, wait = self.setup_driver()
            
            # Navigate to the application status page
            driver.get(APPLICATION_STATUS_URL)

            # Handle any alerts
            try:
                alert = driver.switch_to.alert
                alert.accept()
            except NoAlertPresentException:
                pass

            # Input application number
            element1 = wait.until(EC.element_to_be_clickable((By.ID, "ApplicationNumber")))
            element1.clear()
            element1.send_keys(application_number)
            time.sleep(1)  # Increased delay for stability

            # Handle CAPTCHA
            captcha_text = self.get_captcha_text(driver, wait)
            input_field1 = wait.until(EC.element_to_be_clickable((By.ID, "CaptchaText")))
            input_field1.clear()
            input_field1.send_keys(captcha_text)
            time.sleep(1)  # Increased delay for stability

            # Submit form
            submit_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//input[@value='Show Status']")))
            submit_button.click()
            time.sleep(2)  # Wait for response

            # Extract data
            data = self.extract_application_data(driver, wait, application_number)
            if data is None and retry_count < MAX_RETRIES:
                time.sleep(RETRY_DELAY)
                if driver:
                    driver.quit()
                return self.process_application_number(application_number, retry_count + 1)
            return application_number, data

        except Exception as e:
            st.error(f"Error processing application {application_number}: {str(e)}")
            if retry_count < MAX_RETRIES:
                time.sleep(RETRY_DELAY)
                if driver:
                    driver.quit()
                return self.process_application_number(application_number, retry_count + 1)
            return application_number, None
        finally:
            if driver:
                driver.quit()

    def get_captcha_text(self, driver, wait):
        """Retrieve and parse CAPTCHA text."""
        driver.execute_script(f"window.open('{CAPTCHA_URL}','_blank');")
        driver.switch_to.window(driver.window_handles[-1])
        
        element2 = wait.until(EC.presence_of_element_located((By.TAG_NAME, "pre")))
        captcha = element2.text
        json_data = json.loads(captcha)
        
        driver.close()
        driver.switch_to.window(driver.window_handles[0])
        
        return json_data["CaptchaImageText"]

    def extract_application_data(self, driver, wait, application_number):
        """Extract application data from the page."""
        try:
            body = wait.until(EC.presence_of_element_located((By.TAG_NAME, "body")))
            tables = body.find_elements(By.TAG_NAME, "table")

            if tables:
                data = {"Application Number": application_number}
                for table in tables[:2]:
                    self.extract_table_data(table, data)
                return data
            return None

        except TimeoutException:
            return None
        except Exception as e:
            st.error(f"Error extracting data: {str(e)}")
            return None

    def extract_table_data(self, table, data):
        """Extract data from table rows."""
        rows = table.find_elements(By.TAG_NAME, "tr")
        for row in rows:
            columns = row.find_elements(By.TAG_NAME, "td")
            if len(columns) == 2:
                key = columns[0].text.strip()
                value = columns[1].text.strip()
                if key == "APPLICANT NAME":
                    data["Applicant Name"] = value
                elif key == "APPLICATION TYPE":
                    data["Application Type"] = value
                elif key == "DATE OF FILING":
                    data["Date of Filing"] = parse_date(value)
                elif key == "TITLE OF INVENTION":
                    data["Title of Invention"] = value
                elif key == "FIELD OF INVENTION":
                    data["Field of Invention"] = value
                elif key == "E-MAIL (As Per Record)":
                    data["Email (As Per Record)"] = value
                elif key == "ADDITIONAL-EMAIL (As Per Record)":
                    data["Additional Email (As Per Record)"] = value
                elif key == "E-MAIL (UPDATED Online)":
                    data["Email (Updated Online)"] = value
                elif key == "PCT INTERNATIONAL APPLICATION NUMBER":
                    data["PCT International Application Number"] = value
                elif key == "PCT INTERNATIONAL FILING DATE":
                    data["PCT International Filing Date"] = parse_date(value)
                elif key == "PRIORITY DATE":
                    data["Priority Date"] = parse_date(value)
                elif key == "REQUEST FOR EXAMINATION DATE":
                    data["Request for Examination Date"] = parse_date(value)
                elif key == "PUBLICATION DATE (U/S 11A)":
                    data["Publication Date (U/S 11A)"] = parse_date(value)
                elif key == "APPLICATION STATUS":
                    data["Application Status"] = value

    def export_to_excel(self, data, file_path):
        """Export processed data to Excel file."""
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Application Statuses"

        headers = [
            "Application Number", "Applicant Name", "Application Type", "Date of Filing",
            "Title of Invention", "Field of Invention", "Email (As Per Record)",
            "Additional Email (As Per Record)", "Email (Updated Online)",
            "PCT International Application Number", "PCT International Filing Date",
            "Priority Date", "Request for Examination Date", "Publication Date (U/S 11A)",
            "Application Status"
        ]
        sheet.append(headers)

        for application_number, app_data in data:
            if app_data:
                row = [
                    app_data.get("Application Number", ""),
                    app_data.get("Applicant Name", ""),
                    app_data.get("Application Type", ""),
                    app_data.get("Date of Filing", ""),
                    app_data.get("Title of Invention", ""),
                    app_data.get("Field of Invention", ""),
                    app_data.get("Email (As Per Record)", ""),
                    app_data.get("Additional Email (As Per Record)", ""),
                    app_data.get("Email (Updated Online)", ""),
                    app_data.get("PCT International Application Number", ""),
                    app_data.get("PCT International Filing Date", ""),
                    app_data.get("Priority Date", ""),
                    app_data.get("Request for Examination Date", ""),
                    app_data.get("Publication Date (U/S 11A)", ""),
                    app_data.get("Application Status", "")
                ]
                sheet.append(row)

        workbook.save(file_path)

def main():
    st.title("📋 Patent Application Status Checker")
    st.write("Upload an Excel file containing patent application numbers to check their status.")

    # File uploader
    uploaded_file = st.file_uploader("Upload Excel File", type="xlsx", 
                                   help="Upload an Excel file containing application numbers in the first column")
    
    if uploaded_file:
        try:
            # Save uploaded file
            with open("temp.xlsx", "wb") as f:
                f.write(uploaded_file.getbuffer())

            if st.button("Start Processing", type="primary"):
                with st.spinner("Processing applications..."):
                    service = ApplicationService()
                    workbook = load_workbook(filename="temp.xlsx")
                    sheet = workbook.active
                    application_numbers = [str(row[0]) for row in sheet.iter_rows(min_row=2, values_only=True) if row[0]]

                    if not application_numbers:
                        st.error("No application numbers found in the Excel file.")
                        return

                    st.info(f"Found {len(application_numbers)} application numbers to process.")
                    progress_bar = st.progress(0)
                    results = []

                    # Process each application number
                    for idx, app_num in enumerate(application_numbers):
                        try:
                            result = service.process_application_number(app_num)
                            results.append(result)
                            progress_bar.progress((idx + 1) / len(application_numbers))
                        except Exception as e:
                            st.error(f"Error processing {app_num}: {str(e)}")
                            results.append((app_num, None))

                    # Export results
                    output_file = "output.xlsx"
                    service.export_to_excel(results, output_file)

                    # Provide download button
                    with open(output_file, "rb") as f:
                        st.success("✅ Processing completed successfully!")
                        st.download_button(
                            label="📥 Download Results",
                            data=f,
                            file_name="patent_status_results.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )

        except Exception as e:
            st.error(f"An error occurred: {str(e)}")
            st.write("Please try again or contact support if the error persists.")

        finally:
            # Clean up temporary files
            if os.path.exists("temp.xlsx"):
                os.remove("temp.xlsx")
            if os.path.exists("output.xlsx"):
                os.remove("output.xlsx")

    # Add instructions
    with st.expander("📖 Instructions"):
        st.write("""
        1. Prepare an Excel file with application numbers in the first column (starting from row 2)
        2. Upload the Excel file using the file uploader above
        3. Click 'Start Processing' to begin checking the status
        4. Wait for processing to complete
        5. Download the results file when ready
        
        Note: The process might take a few minutes depending on the number of applications.
        """)

if __name__ == "__main__":
    main()